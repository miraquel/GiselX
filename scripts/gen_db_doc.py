"""Generate GiselX database documentation workbook (tables + fields + SPs)."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = r"D:\Programming\GiselX\GiselX-Database-Documentation.xlsx"

# ---- styling helpers -------------------------------------------------------
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14, color="1F4E78")
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(vertical="top", wrap_text=True)
TOP = Alignment(vertical="top")

GREEN = PatternFill("solid", fgColor="E2EFDA")   # EF-managed
RED = PatternFill("solid", fgColor="FCE4D6")     # no DDL / missing
PK_FILL = PatternFill("solid", fgColor="FFF2CC") # key columns

def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = BORDER

def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ---- table field data ------------------------------------------------------
# (schema, table, creation, [ (col, type, nullable, key/notes) ... ])
N = "NULL"
NN = "NOT NULL"

tables = [
    ("identity", "AspNetUsers", "EF migration", [
        ("Id", "nvarchar(450)", NN, "PK"),
        ("UserName", "nvarchar(256)", N, ""),
        ("NormalizedUserName", "nvarchar(256)", N, "Unique index UserNameIndex"),
        ("Email", "nvarchar(256)", N, ""),
        ("NormalizedEmail", "nvarchar(256)", N, "Index EmailIndex"),
        ("EmailConfirmed", "bit", NN, ""),
        ("PasswordHash", "nvarchar(max)", N, ""),
        ("SecurityStamp", "nvarchar(max)", N, ""),
        ("ConcurrencyStamp", "nvarchar(max)", N, "Concurrency token"),
        ("PhoneNumber", "nvarchar(max)", N, ""),
        ("PhoneNumberConfirmed", "bit", NN, ""),
        ("TwoFactorEnabled", "bit", NN, ""),
        ("LockoutEnd", "datetimeoffset", N, ""),
        ("LockoutEnabled", "bit", NN, ""),
        ("AccessFailedCount", "int", NN, ""),
        ("CompanyId", "int", NN, "FK -> dbo.Company (custom); index"),
        ("FirstName", "nvarchar(max)", NN, "Custom field"),
        ("LastName", "nvarchar(max)", N, "Custom field"),
    ]),
    ("identity", "AspNetRoles", "EF migration", [
        ("Id", "nvarchar(450)", NN, "PK"),
        ("Name", "nvarchar(256)", N, ""),
        ("NormalizedName", "nvarchar(256)", N, "Unique index RoleNameIndex"),
        ("ConcurrencyStamp", "nvarchar(max)", N, "Concurrency token"),
        ("Description", "nvarchar(max)", NN, "Custom field"),
    ]),
    ("identity", "AspNetUserRoles", "EF migration", [
        ("UserId", "nvarchar(450)", NN, "PK; FK -> AspNetUsers"),
        ("RoleId", "nvarchar(450)", NN, "PK; FK -> AspNetRoles; index"),
    ]),
    ("identity", "AspNetUserClaims", "EF migration", [
        ("Id", "int", NN, "PK; identity"),
        ("UserId", "nvarchar(450)", NN, "FK -> AspNetUsers; index"),
        ("ClaimType", "nvarchar(max)", N, ""),
        ("ClaimValue", "nvarchar(max)", N, ""),
    ]),
    ("identity", "AspNetUserLogins", "EF migration", [
        ("LoginProvider", "nvarchar(128)", NN, "PK"),
        ("ProviderKey", "nvarchar(128)", NN, "PK"),
        ("ProviderDisplayName", "nvarchar(max)", N, ""),
        ("UserId", "nvarchar(450)", NN, "FK -> AspNetUsers; index"),
    ]),
    ("identity", "AspNetRoleClaims", "EF migration", [
        ("Id", "int", NN, "PK; identity"),
        ("RoleId", "nvarchar(450)", NN, "FK -> AspNetRoles; index"),
        ("ClaimType", "nvarchar(max)", N, ""),
        ("ClaimValue", "nvarchar(max)", N, ""),
    ]),
    ("identity", "AspNetUserTokens", "EF migration", [
        ("UserId", "nvarchar(450)", NN, "PK; FK -> AspNetUsers"),
        ("LoginProvider", "nvarchar(128)", NN, "PK"),
        ("Name", "nvarchar(128)", NN, "PK"),
        ("Value", "nvarchar(max)", N, ""),
    ]),
    ("dbo", "Company", "EF migration", [
        ("Id", "int", NN, "PK; identity"),
        ("Name", "nvarchar(100)", NN, "Unique index"),
        ("Address", "nvarchar(255)", N, ""),
        ("ContactEmail", "nvarchar(255)", N, "Reminder field"),
        ("ReminderLeadDays", "int", NN, "Default 3"),
        ("DeadlineDaysOfWeek", "int", N, "Flags enum (WeekDays)"),
        ("DeadlineDayOfMonth", "int", N, ""),
    ]),
    ("dbo", "TransDist", "EF migration", [
        ("Id", "int", NN, "PK; identity"),
        ("SoId", "nvarchar(20)", N, ""),
        ("SoCreateDate", "datetime", N, ""),
        ("LeadTimeDlv", "int", N, ""),
        ("LeadTimeRct", "int", N, ""),
        ("ItemId", "nvarchar(20)", N, ""),
        ("ItemName", "nvarchar(70)", N, ""),
        ("SoQty", "numeric(32,16)", N, ""),
        ("Unit", "nvarchar(20)", N, ""),
        ("KgPerUnit", "numeric(32,16)", N, ""),
        ("DlvDateRequest", "datetime", N, ""),
        ("RctDateRequest", "datetime", N, ""),
        ("DoQty", "numeric(32,16)", N, ""),
        ("DoDate", "datetime", N, ""),
        ("ReceiptDate", "datetime", N, ""),
        ("CreatedBy", "nvarchar(60)", N, ""),
        ("CreatedDate", "datetime", N, ""),
        ("CompanyId", "int", NN, "FK -> dbo.Company (Restrict); index"),
    ]),
    ("dbo", "ServelDeliveryEx", "EF migration", [
        ("Id", "int", NN, "PK; identity"),
        ("SoId", "nvarchar(20)", N, ""),
        ("SoCreateDate", "datetime", N, ""),
        ("ItemId", "nvarchar(20)", N, ""),
        ("ItemName", "nvarchar(4000)", N, ""),
        ("QuadranServel", "nvarchar(20)", N, ""),
        ("SoQty", "numeric(32,16)", N, ""),
        ("SoQtyCalc", "numeric(32,16)", N, ""),
        ("DoQty", "numeric(32,16)", N, ""),
        ("DoQtyCalc", "numeric(32,16)", N, ""),
        ("DoDate", "datetime", N, ""),
        ("DoDateMin", "datetime", N, ""),
        ("ShippingDateRequest", "datetime", N, ""),
        ("ShippingDateRequestCalc", "datetime", N, ""),
        ("ReceiptDate", "datetime", N, ""),
        ("ReceiptDateRequest", "datetime", N, ""),
        ("ReceiptDateRequestCalc", "datetime", N, ""),
        ("LeadTimeDelivery", "int", N, ""),
        ("LeadTimeReceipt", "int", N, ""),
        ("Infull", "int", N, ""),
        ("Ontime", "int", N, ""),
        ("PeriodDate", "date", N, ""),
    ]),
    ("dbo", "ServelReceiptEx", "EF migration", [
        ("Id", "int", NN, "PK; identity"),
        ("SoId", "nvarchar(20)", N, ""),
        ("SoCreateDate", "datetime", N, ""),
        ("ItemId", "nvarchar(20)", N, ""),
        ("ItemName", "nvarchar(4000)", N, ""),
        ("QuadranServel", "nvarchar(20)", N, ""),
        ("SoQty", "numeric(32,16)", N, ""),
        ("SoQtyCalc", "numeric(32,16)", N, ""),
        ("DoQty", "numeric(32,16)", N, ""),
        ("DoQtyCalc", "numeric(32,16)", N, ""),
        ("DoDate", "datetime", N, ""),
        ("DoDateMin", "datetime", N, ""),
        ("ShippingDateRequest", "datetime", N, ""),
        ("ShippingDateRequestCalc", "datetime", N, ""),
        ("ReceiptDate", "datetime", N, ""),
        ("ReceiptDateRequest", "datetime", N, ""),
        ("ReceiptDateRequestCalc", "datetime", N, ""),
        ("LeadTimeDelivery", "int", N, ""),
        ("LeadtimeReceipt", "int", N, "Note: spelling differs from DeliveryEx"),
        ("Infull", "int", N, ""),
        ("Ontime", "int", N, ""),
        ("PeriodDate", "date", N, ""),
    ]),
    ("dbo", "SalesTransaction", "NO DDL (inferred)", [
        ("Id", "int", NN, "PK; identity (assumed)"),
        ("CustomerId", "nvarchar", "?", "Inferred from CLR string"),
        ("CustomerName", "nvarchar", "?", ""),
        ("CustomerAlias", "nvarchar", "?", ""),
        ("CustomerAddress", "nvarchar", "?", ""),
        ("ProductId", "nvarchar", "?", ""),
        ("ProductBrand", "nvarchar", "?", ""),
        ("ProductName", "nvarchar", "?", ""),
        ("ProductPackaging", "nvarchar", "?", ""),
        ("ProductPcsInCtn", "int", "?", ""),
        ("BatchNumber", "nvarchar", "?", ""),
        ("ExpiredDate", "datetime", "?", ""),
        ("InvoiceNo", "nvarchar", "?", ""),
        ("InvoiceDate", "datetime", "?", "Period filter source"),
        ("InvoiceQty", "decimal", "?", ""),
        ("InvoiceUnit", "nvarchar", "?", ""),
        ("UnitPrice", "decimal", "?", ""),
        ("GrossValue", "decimal", "?", ""),
        ("DiscountValue", "decimal", "?", ""),
        ("DiscountPct", "decimal", "?", ""),
        ("NetValue", "decimal", "?", ""),
        ("InKg", "decimal", "?", ""),
        ("GeisaPOId", "nvarchar", "?", ""),
        ("ShipDate", "datetime", "?", ""),
        ("SalesmanNameGMK", "nvarchar", "?", ""),
        ("CompanyId", "int", "?", "Filter key"),
        ("CreatedDate", "datetime", "?", "Latest-upload-batch key"),
    ]),
    ("dbo", "Stock", "NO DDL (inferred)", [
        ("Id", "int", NN, "PK; identity (assumed)"),
        ("ProductId", "nvarchar", "?", "Inferred from CLR string"),
        ("ProductName", "nvarchar", "?", ""),
        ("ProductPackaging", "nvarchar", "?", ""),
        ("ProductPcsInCtn", "int", "?", ""),
        ("ProductNetto", "decimal", "?", ""),
        ("ProductUnit", "nvarchar", "?", ""),
        ("SaldoAwal", "decimal", "?", ""),
        ("SaldoMasukPO", "decimal", "?", ""),
        ("SaldoAkhir", "decimal", "?", ""),
        ("BatchNumber", "nvarchar", "?", ""),
        ("ExpiredDate", "datetime", "?", ""),
        ("CompanyId", "int", "?", "Filter key"),
        ("CreatedDate", "datetime", "?", "Latest-upload-batch key"),
    ]),
]

sps = [
    ("Company_Insert", "dbo.Company", "INSERT", "Yes", ""),
    ("Company_Update", "dbo.Company", "UPDATE", "Yes", ""),
    ("Company_SelectById", "dbo.Company", "SELECT", "Yes", ""),
    ("Company_SelectForReminder", "dbo.Company", "SELECT", "Yes", "Used by reminder job"),
    ("Company_Select", "dbo.Company", "SELECT", "NO - MISSING", "Called CompanyRepository.cs:65"),
    ("Gisel_Select", "dbo.TransDist", "SELECT (paged)", "Yes", ""),
    ("Gisel_SelectByCust", "dbo.TransDist", "SELECT", "Yes", ""),
    ("Gisel_SelectByCustPeriod", "dbo.TransDist", "SELECT", "Yes", ""),
    ("Gisel_SelectPeriods", "dbo.TransDist", "SELECT", "Yes", ""),
    ("Gisel_del", "dbo.TransDist", "DELETE", "Yes", ""),
    ("GiselStock_Select", "dbo.Stock", "SELECT (paged)", "Yes", ""),
    ("GiselStock_SelectByCust", "dbo.Stock", "SELECT", "Yes", ""),
    ("GiselStock_SelectByCustPeriod", "dbo.Stock", "SELECT", "Yes", ""),
    ("GiselStock_SelectPeriods", "dbo.Stock", "SELECT", "Yes", ""),
    ("GiselSales_Select", "dbo.SalesTransaction", "SELECT (paged)", "Yes", ""),
    ("GiselSales_SelectByCust", "dbo.SalesTransaction", "SELECT", "Yes", ""),
    ("GiselSales_SelectByCustPeriod", "dbo.SalesTransaction", "SELECT", "Yes", ""),
    ("GiselSales_SelectPeriods", "dbo.SalesTransaction", "SELECT", "Yes", ""),
    ("GiselSales_del", "dbo.SalesTransaction", "DELETE", "NO - MISSING", "Called SalesTransactionRepository.cs:143"),
]

# ---- build workbook --------------------------------------------------------
wb = openpyxl.Workbook()

# ===== Sheet 1: Overview =====
ws = wb.active
ws.title = "Overview"
ws["A1"] = "GiselX Database Documentation"
ws["A1"].font = TITLE_FONT
ws["A2"] = "Reverse-engineered from codebase (EF migrations, model snapshot, domain entities, stored procedures)."
ws["A2"].font = Font(italic=True, color="808080")
ws["A3"] = "Generated for production-database rebuild. 13 tables, 19 stored procedures."
ws["A3"].font = Font(italic=True, color="808080")

hdr = ["Schema", "Table", "Creation method", "# Fields", "Source", "Notes"]
ws.append([])
hrow = 5
for i, h in enumerate(hdr, start=1):
    ws.cell(row=hrow, column=i, value=h)
style_header(ws, hrow, len(hdr))

src_map = {
    "EF migration": "Migrations + model snapshot",
    "NO DDL (inferred)": "Domain entity + SPs only",
}
note_map = {
    "NO DDL (inferred)": "No CREATE TABLE anywhere - types/lengths inferred",
}
r = hrow + 1
for schema, table, creation, cols in tables:
    ws.cell(row=r, column=1, value=schema)
    ws.cell(row=r, column=2, value=table)
    ws.cell(row=r, column=3, value=creation)
    ws.cell(row=r, column=4, value=len(cols))
    ws.cell(row=r, column=5, value=src_map.get(creation, ""))
    ws.cell(row=r, column=6, value=note_map.get(creation, ""))
    fill = RED if creation.startswith("NO DDL") else GREEN
    for c in range(1, len(hdr) + 1):
        cell = ws.cell(row=r, column=c)
        cell.fill = fill
        cell.border = BORDER
        cell.alignment = TOP
    r += 1
autosize(ws, [12, 22, 20, 9, 30, 48])
ws.freeze_panes = "A6"
ws.auto_filter.ref = f"A{hrow}:F{r-1}"

# ===== Sheet 2: Columns (flat) =====
ws2 = wb.create_sheet("Columns")
hdr2 = ["Schema", "Table", "Field", "Data Type", "Nullable", "Key / Notes"]
for i, h in enumerate(hdr2, start=1):
    ws2.cell(row=1, column=i, value=h)
style_header(ws2, 1, len(hdr2))
r = 2
for schema, table, creation, cols in tables:
    for (col, typ, nullable, note) in cols:
        ws2.cell(row=r, column=1, value=schema)
        ws2.cell(row=r, column=2, value=table)
        ws2.cell(row=r, column=3, value=col)
        ws2.cell(row=r, column=4, value=typ)
        ws2.cell(row=r, column=5, value=nullable)
        ws2.cell(row=r, column=6, value=note)
        is_key = "PK" in note or "FK" in note
        for c in range(1, len(hdr2) + 1):
            cell = ws2.cell(row=r, column=c)
            cell.border = BORDER
            cell.alignment = WRAP if c == 6 else TOP
            if creation.startswith("NO DDL"):
                cell.fill = RED
            elif is_key:
                cell.fill = PK_FILL
        r += 1
autosize(ws2, [12, 20, 26, 18, 12, 44])
ws2.freeze_panes = "A2"
ws2.auto_filter.ref = f"A1:F{r-1}"

# ===== Sheet 3: one block per table (readable layout) =====
ws3 = wb.create_sheet("By Table")
r = 1
for schema, table, creation, cols in tables:
    tcell = ws3.cell(row=r, column=1, value=f"{schema}.{table}")
    tcell.font = Font(bold=True, size=12, color="FFFFFF")
    badge = ws3.cell(row=r, column=4, value=creation)
    badge.font = Font(bold=True, italic=True,
                      color="C00000" if creation.startswith("NO DDL") else "375623")
    for c in range(1, 5):
        ws3.cell(row=r, column=c).fill = HEADER_FILL
    r += 1
    for i, h in enumerate(["Field", "Data Type", "Nullable", "Key / Notes"], start=1):
        hc = ws3.cell(row=r, column=i, value=h)
        hc.font = Font(bold=True)
        hc.fill = PatternFill("solid", fgColor="D6DCE4")
        hc.border = BORDER
    r += 1
    for (col, typ, nullable, note) in cols:
        ws3.cell(row=r, column=1, value=col)
        ws3.cell(row=r, column=2, value=typ)
        ws3.cell(row=r, column=3, value=nullable)
        ws3.cell(row=r, column=4, value=note)
        for c in range(1, 5):
            cell = ws3.cell(row=r, column=c)
            cell.border = BORDER
            cell.alignment = WRAP if c == 4 else TOP
            if "PK" in note or "FK" in note:
                cell.fill = PK_FILL
        r += 1
    r += 1  # blank spacer row
autosize(ws3, [26, 18, 12, 50])

# ===== Sheet 4: Stored Procedures =====
ws4 = wb.create_sheet("Stored Procedures")
hdr4 = ["Stored Procedure", "Primary Table", "Operation", "Script in repo?", "Notes"]
for i, h in enumerate(hdr4, start=1):
    ws4.cell(row=1, column=i, value=h)
style_header(ws4, 1, len(hdr4))
r = 2
for name, table, op, present, note in sps:
    ws4.cell(row=r, column=1, value=name)
    ws4.cell(row=r, column=2, value=table)
    ws4.cell(row=r, column=3, value=op)
    ws4.cell(row=r, column=4, value=present)
    ws4.cell(row=r, column=5, value=note)
    missing = present.startswith("NO")
    for c in range(1, len(hdr4) + 1):
        cell = ws4.cell(row=r, column=c)
        cell.border = BORDER
        cell.alignment = TOP
        if missing:
            cell.fill = RED
    r += 1
autosize(ws4, [30, 22, 16, 16, 40])
ws4.freeze_panes = "A2"
ws4.auto_filter.ref = f"A1:E{r-1}"

# ===== Sheet 5: Legend =====
ws5 = wb.create_sheet("Legend")
ws5["A1"] = "Legend & Rebuild Notes"
ws5["A1"].font = TITLE_FONT
legend = [
    ("", ""),
    ("Green rows", "Table created by EF Core migrations - run 'dotnet ef database update' to recreate."),
    ("Orange/red rows", "Table or SP has NO definition in the repo - must be authored by hand to rebuild."),
    ("Yellow cells", "Primary key / foreign key columns."),
    ("", ""),
    ("Rebuild gap 1", "dbo.SalesTransaction - no CREATE TABLE; columns/types inferred from SalesTransaction.cs."),
    ("Rebuild gap 2", "dbo.Stock - no CREATE TABLE; columns/types inferred from Stock.cs."),
    ("Rebuild gap 3", "Company_Select SP - called in code but no .sql script exists."),
    ("Rebuild gap 4", "GiselSales_del SP - called in code but no .sql script exists."),
    ("", ""),
    ("Type note", "'?' nullable = unknown (no DDL). nvarchar/decimal lengths/precision for Stock & "
                  "SalesTransaction are not defined in the repo."),
    ("SP families", "Gisel_* -> TransDist | GiselStock_* -> Stock | GiselSales_* -> SalesTransaction | "
                    "Company_* -> Company"),
]
r = 3
for a, b in legend:
    ws5.cell(row=r, column=1, value=a).font = Font(bold=True)
    ws5.cell(row=r, column=2, value=b).alignment = WRAP
    r += 1
autosize(ws5, [18, 95])

wb.save(OUT)
print("Saved:", OUT)
print("Sheets:", wb.sheetnames)
print("Tables:", len(tables), "| Total columns:", sum(len(t[3]) for t in tables), "| SPs:", len(sps))
